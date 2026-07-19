#!/usr/bin/env python3
"""
Conversor del Excel de la porra -> docs/datos.json

Lee el .xlsx (valores ya calculados por las fórmulas de Excel) y genera un JSON
que consume la web. NO recalcula puntos: lee los que el propio Excel ha calculado,
así la web siempre cuadra con tu hoja.

Uso:
    python3 exportar.py                # usa el .xlsx por defecto (carpeta superior)
    python3 exportar.py ruta/al.xlsx   # ruta concreta
"""
import sys, os, json, datetime, re, unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala con:  pip3 install openpyxl")

# ---------------------------------------------------------------- rutas
AQUI = os.path.dirname(os.path.abspath(__file__))
DEFECTO_XLSX = os.path.join(AQUI, "..", "ADMIN lMundial 2026 - Porra Poyete.xlsx")
SALIDA = os.path.join(AQUI, "docs", "datos.json")
EXTRAS = os.path.join(AQUI, "resultados_extra.json")

# ---------------------------------------------------------------- banderas
BANDERAS = {
    "México":"🇲🇽","Sudáfrica":"🇿🇦","Corea del Sur":"🇰🇷","República Checa":"🇨🇿",
    "Canadá":"🇨🇦","Bosnia y Herzegovina":"🇧🇦","Catar":"🇶🇦","Suiza":"🇨🇭",
    "Brasil":"🇧🇷","Marruecos":"🇲🇦","Escocia":"🏴󠁧󠁢󠁳󠁣󠁴󠁿","Haití":"🇭🇹",
    "Estados Unidos":"🇺🇸","Paraguay":"🇵🇾","Australia":"🇦🇺","Turquía":"🇹🇷",
    "Alemania":"🇩🇪","Curazao":"🇨🇼","Costa de Marfil":"🇨🇮","Ecuador":"🇪🇨",
    "Países Bajos":"🇳🇱","Japón":"🇯🇵","Suecia":"🇸🇪","Túnez":"🇹🇳",
    "Bélgica":"🇧🇪","Egipto":"🇪🇬","Irán":"🇮🇷","Nueva Zelanda":"🇳🇿",
    "España":"🇪🇸","Cabo Verde":"🇨🇻","Arabia Saudita":"🇸🇦","Uruguay":"🇺🇾",
    "Francia":"🇫🇷","Senegal":"🇸🇳","Irak":"🇮🇶","Noruega":"🇳🇴",
    "Argentina":"🇦🇷","Argelia":"🇩🇿","Austria":"🇦🇹","Jordania":"🇯🇴",
    "Portugal":"🇵🇹","RD Congo":"🇨🇩","Uzbekistán":"🇺🇿","Colombia":"🇨🇴",
    "Inglaterra":"🏴󠁧󠁢󠁥󠁮󠁧󠁿","Croacia":"🇭🇷","Ghana":"🇬🇭","Panamá":"🇵🇦",
}
def bandera(equipo):
    return BANDERAS.get((equipo or "").strip(), "")

# ---------------------------------------------------------------- helpers
def col_letra(n):
    return openpyxl.utils.get_column_letter(n)

def parse_pred(valor):
    """ '1|2-0'  ->  {'signo':'1','local':2,'visitante':0}
        'Corea del Sur-Canadá·X|1-1' -> idem + 'duelo':'Corea del Sur-Canadá'
        '-' o vacío -> None """
    if valor is None:
        return None
    s = str(valor).strip()
    if s in ("-", "", "Pendiente"):
        return None
    duelo = None
    if "·" in s:
        duelo, s = s.split("·", 1)
    if "|" not in s:
        # predicción de equipo (clasificados, cuadro de honor) -> texto plano
        return {"texto": valor}
    signo, marcador = s.split("|", 1)
    m = re.match(r"\s*(\d+)\s*-\s*(\d+)\s*", marcador)
    if not m:
        return {"signo": signo.strip(), "texto": valor}
    out = {"signo": signo.strip(), "local": int(m.group(1)), "visitante": int(m.group(2))}
    if duelo:
        out["duelo"] = duelo.strip()
    return out

def parse_marcador(valor, signo=None):
    """Resultado real en formato del Excel: '1|2-0' en grupos o '2-1' en KO."""
    pred = parse_pred(valor)
    if pred and "local" in pred:
        return pred

    if valor is None:
        return None
    s = str(valor).strip()
    if s in ("-", "", "Pendiente"):
        return None
    m = re.match(r"\s*(\d+)\s*-\s*(\d+)\s*", s)
    if not m:
        return None

    local = int(m.group(1))
    visitante = int(m.group(2))
    signo = str(signo or "").strip()
    if signo == "1":
        signo = "1"
    elif signo == "2":
        signo = "2"
    elif signo in ("0", "X"):
        signo = "X"
    elif local > visitante:
        signo = "1"
    elif visitante > local:
        signo = "2"
    else:
        signo = "X"
    return {"signo": signo, "local": local, "visitante": visitante}

def iso_fecha(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return None

def compacta(texto):
    return re.sub(r"\s+", " ", str(texto or "")).strip()

def normaliza_nombre(texto):
    texto = unicodedata.normalize("NFKD", str(texto or "").lower())
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9 ]+", " ", texto)
    texto = re.sub(r"\b(jr|junior)\b", "", texto)
    return compacta(texto)

def acierta_honor(prediccion, resultado):
    pred = normaliza_nombre(prediccion)
    real = normaliza_nombre(resultado)
    if not pred or not real:
        return False
    return pred == real or pred in real or real in pred

# ---------------------------------------------------------------- carga
ruta = sys.argv[1] if len(sys.argv) > 1 else DEFECTO_XLSX
if not os.path.exists(ruta):
    sys.exit(f"No encuentro el Excel en: {ruta}")

if os.path.exists(EXTRAS):
    with open(EXTRAS, encoding="utf-8") as fh:
        resultados_extra = json.load(fh)
else:
    resultados_extra = {}

etiquetas = resultados_extra.get("_etiquetas", {})
cuadro_honor_extra = {
    compacta(k): v for k, v in resultados_extra.get("_cuadro_honor", {}).items()
}

def sustituye_etiquetas(valor):
    s = str(valor or "")
    for etiqueta, equipo in etiquetas.items():
        s = re.sub(rf"\b{re.escape(etiqueta)}\b", str(equipo), s)
    return s

import warnings
warnings.simplefilter("ignore")
wb = openpyxl.load_workbook(ruta, data_only=True)
adm = wb["ADMIN"]
clas = wb["CLAS"]
home = wb["Home"]

# Hora de Madrid por partido, desde la hoja Horarios (col C = hora Madrid)
hora_madrid = {}  # (fecha_str, grupo_letter) -> "HH:MM"
if "Horarios" in wb.sheetnames:
    hor = wb["Horarios"]
    cur_g = None
    for r in range(1, hor.max_row + 1):
        v = hor.cell(r, 1).value
        if isinstance(v, str) and len(v) == 2 and v[0] == "G":
            cur_g = v[1]
        elif isinstance(v, datetime.datetime) and cur_g:
            madrid = hor.cell(r, 3).value
            if isinstance(madrid, datetime.datetime):
                key = (madrid.strftime("%Y-%m-%d"), cur_g)
                if key not in hora_madrid:
                    hora_madrid[key] = madrid.strftime("%H:%M")

# ---------------------------------------------------------------- jugadores
# Nombres en ADMIN fila 5, columnas 19,22,25,... (cada 3). La predicción va en
# esa columna y los puntos en la de al lado (col+1).
jugadores = []      # [{'nombre':..,'col_pred':..,'col_pts':..}]
col = 19
while col <= 70:
    v = adm.cell(5, col).value
    if isinstance(v, str) and v.strip() and not v.strip().isdigit():
        jugadores.append({"nombre": v.strip(), "col_pred": col, "col_pts": col + 1})
    col += 3
nombres = [j["nombre"] for j in jugadores]

# ---------------------------------------------------------------- partidos
# Filas de partido = aquellas cuya predicción tiene formato 'signo|marcador'.
# Vamos arrastrando la cabecera de fase (filas en mayúsculas en col K).
FASES = {
    "FASE DE GRUPOS": "Fase de grupos",
    "ENFRENTAMIENTOS DIECISEISAVOS": "Dieciseisavos",
    "ENFRENTAMIENTOS OCTAVOS": "Octavos",
    "ENFRENTAMIENTOS CUARTOS": "Cuartos",
    "ENFRENTAMIENTOS SEMIFINALES": "Semifinales",
    "3º-4º PUESTO": "3º y 4º puesto",
    "ENFRENTAMIENTO FINAL": "Final",
}
partidos = []
fase_actual = "Fase de grupos"
for r in range(6, 260):
    k = adm.cell(r, 11).value          # nombre del partido / cabecera
    if isinstance(k, str) and k.strip().upper() in FASES:
        fase_actual = FASES[k.strip().upper()]
        continue
    if not isinstance(k, str) or "-" not in k:
        continue
    # ¿es un partido jugable? -> al menos una predicción con '|'
    preds_raw = {j["nombre"]: adm.cell(r, j["col_pred"]).value for j in jugadores}
    if not any(isinstance(v, str) and "|" in v for v in preds_raw.values()):
        continue

    fecha = iso_fecha(adm.cell(r, 8).value)     # col H
    jcod = adm.cell(r, 10).value                # col J (A1, 1/16, ...)
    grupo, jornada = None, None
    if isinstance(jcod, str) and re.match(r"^[A-L]\d$", jcod):
        grupo, jornada = jcod[0], "J" + jcod[1]

    # equipos: K = "Local-Visitante"; en rondas futuras puede contener W73-W74.
    codigo_original = k.strip()
    codigo_visible = sustituye_etiquetas(codigo_original)
    equipos = None
    partes = [p.strip() for p in codigo_visible.split("-", 1)]
    if len(partes) == 2 and not any(re.match(r"^[WL]\d+$", p) for p in partes):
        equipos = partes

    resultado = parse_marcador(adm.cell(r, 13).value, adm.cell(r, 12).value)   # col M, signo col L
    if resultado:
        resultado.update(resultados_extra.get(codigo_original, resultados_extra.get(codigo_visible, {})))
    jugado = bool(resultado and "local" in resultado)

    preds = {}
    for j in jugadores:
        p = parse_pred(adm.cell(r, j["col_pred"]).value)
        pts = adm.cell(r, j["col_pts"]).value
        if p is None:
            continue
        if isinstance(pts, (int, float)):
            p["puntos"] = round(float(pts), 2)
        preds[j["nombre"]] = p

    hora = hora_madrid.get((fecha, grupo)) if fecha and grupo else None
    partidos.append({
        "id": r,
        "fase": fase_actual,
        "grupo": grupo,
        "jornada": jornada,
        "codigo": codigo_visible,
        "fecha": fecha,
        "hora": hora,
        "equipos": equipos,
        "banderas": [bandera(e) for e in equipos] if equipos else None,
        "resultado": resultado if jugado else None,
        "jugado": jugado,
        "predicciones": preds,
    })

# ---------------------------------------------------------------- cuadro de honor
honor = []
honor_puntos = {nombre: 0.0 for nombre in nombres}
ETIQUETAS_HONOR = {
    "🥇Campeón", "🥈Subcampeón", "🥉3º puesto",
    "Bota de Oro  (máximo goleador)", "Bota de Plata (2º máximo goleador)",
    "Bota de Bronce (3º máximo goleador)", "Balón de Oro  (mejor jugador)",
    "Balón de Plata (2º mejor jugador)", "Balón de Bronce (3º mejor jugador)",
}
PUNTOS_HONOR = {
    "🥇Campeón": 10,
    "🥈Subcampeón": 8,
    "🥉3º puesto": 7,
    "Bota de Oro (máximo goleador)": 8,
    "Bota de Plata (2º máximo goleador)": 5,
    "Bota de Bronce (3º máximo goleador)": 3,
    "Balón de Oro (mejor jugador)": 8,
    "Balón de Plata (2º mejor jugador)": 5,
    "Balón de Bronce (3º mejor jugador)": 3,
}
for r in range(249, 260):
    k = adm.cell(r, 11).value
    if not isinstance(k, str) or k.strip() not in ETIQUETAS_HONOR:
        continue
    concepto = compacta(k)
    real = adm.cell(r, 13).value
    real = real.strip() if isinstance(real, str) else None
    # los placeholders de fórmula (WF, LF, W34, "Escribe el jugador...") = aún sin decidir
    if real and (real.startswith("Escribe") or re.match(r"^[WL]F?\d*$", real)):
        real = None
    real = cuadro_honor_extra.get(concepto, real)
    preds = {}
    for j in jugadores:
        v = adm.cell(r, j["col_pred"]).value
        if isinstance(v, str) and v.strip():
            pts = PUNTOS_HONOR.get(concepto, 0) if acierta_honor(v, real) else 0
            honor_puntos[j["nombre"]] += pts
            preds[j["nombre"]] = {
                "texto": v.strip(),
                "puntos": round(float(pts), 2),
            }
    honor.append({
        "concepto": concepto,
        "resultado": real,
        "bandera": bandera(real),
        "predicciones": preds,
    })

# ---------------------------------------------------------------- clasificación
categorias = []
for c in range(5, 20):                      # E..S en fila 4
    v = clas.cell(4, c).value
    if isinstance(v, str) and v.strip():
        categorias.append({"col": c, "nombre": v.strip()})

clasificacion = []
for r in range(5, 30):
    nombre = clas.cell(r, 3).value          # col C
    if not isinstance(nombre, str) or not nombre.strip() or nombre.strip() == "-":
        continue
    total = clas.cell(r, 4).value           # col D
    if not isinstance(total, (int, float)):
        total = 0
    desglose = {}
    for cat in categorias:
        v = clas.cell(r, cat["col"]).value
        desglose[cat["nombre"]] = float(v) if isinstance(v, (int, float)) else 0
    if "Cuadro de Honor" in desglose:
        desglose["Cuadro de Honor"] = round(honor_puntos.get(nombre.strip(), 0.0), 2)
        total = sum(desglose.values())
    clasificacion.append({
        "jugador": nombre.strip(),
        "total": round(float(total), 2),
        "desglose": desglose,
    })

# ordenar por puntos y asignar posición (con empates)
clasificacion.sort(key=lambda x: (-x["total"], x["jugador"].lower()))
pos = 0
prev = None
for i, fila in enumerate(clasificacion):
    if fila["total"] != prev:
        pos = i + 1
        prev = fila["total"]
    fila["pos"] = pos

# ---------------------------------------------------------------- meta
grupo_nombre = home.cell(10, 3).value or "Porra"
titulo = home.cell(3, 2).value or "Mundial 2026"
partidos_jugados = sum(1 for p in partidos if p["jugado"])

datos = {
    "titulo": str(titulo).strip(),
    "grupo": str(grupo_nombre).strip(),
    "actualizado": datetime.datetime.now().isoformat(timespec="seconds"),
    "jugadores": nombres,
    "categorias": [c["nombre"] for c in categorias],
    "clasificacion": clasificacion,
    "partidos": partidos,
    "cuadro_honor": honor,
    "resumen": {
        "n_jugadores": len(nombres),
        "n_partidos": len(partidos),
        "n_jugados": partidos_jugados,
    },
}

os.makedirs(os.path.dirname(SALIDA), exist_ok=True)
with open(SALIDA, "w", encoding="utf-8") as f:
    json.dump(datos, f, ensure_ascii=False, indent=1)

print(f"✅ Generado {SALIDA}")
print(f"   Jugadores: {len(nombres)}  ->  {', '.join(nombres)}")
print(f"   Partidos:  {len(partidos)}  (jugados: {partidos_jugados})")
print(f"   Cuadro de honor: {len(honor)} categorías")
print(f"   Clasificación líder: {clasificacion[0]['jugador'] if clasificacion else '-'}")
