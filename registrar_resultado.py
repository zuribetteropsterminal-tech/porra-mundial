#!/usr/bin/env python3
"""
Registra un resultado real en el Excel de la porra usando Microsoft Excel.

Uso:
    python3 registrar_resultado.py "Mexico" "Sudafrica" 2 0
    python3 registrar_resultado.py "Mexico-Sudafrica" 2 0

El script localiza el partido en la hoja WORLDCUP, escribe los goles en las
columnas AC/AD, fuerza el calculo del libro y guarda. Despues debe ejecutarse
./actualizar.sh para regenerar datos.json y el mensaje.
"""
import argparse
import os
import subprocess
import sys
import unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala con: pip3 install openpyxl")


AQUI = os.path.dirname(os.path.abspath(__file__))
DEFECTO_XLSX = os.path.abspath(
    os.path.join(AQUI, "..", "ADMIN lMundial 2026 - Porra Poyete.xlsx")
)


def normaliza(texto):
    texto = str(texto or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.replace("–", "-").replace("—", "-")
    return " ".join(texto.split())


def codigo_normalizado(local, visitante):
    return f"{normaliza(local)}-{normaliza(visitante)}"


def partir_codigo(codigo):
    partes = [p.strip() for p in codigo.split("-", 1)]
    if len(partes) != 2 or not partes[0] or not partes[1]:
        raise SystemExit("El partido debe tener formato 'Local-Visitante'.")
    return partes[0], partes[1]


def localizar_partido(ruta, local, visitante):
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["WORLDCUP"]
    objetivo = codigo_normalizado(local, visitante)

    candidatos = []
    for r, _row in enumerate(ws.iter_rows(min_row=4), start=4):
        real_local = ws.cell(r, 27).value  # AA
        real_visitante = ws.cell(r, 32).value  # AF
        codigo = ws.cell(r, 124).value  # DT
        if not real_local or not real_visitante:
            continue

        actual = codigo_normalizado(real_local, real_visitante)
        if actual == objetivo or normaliza(codigo) == objetivo:
            candidatos.append(
                {
                    "fila": r,
                    "local": str(real_local).strip(),
                    "visitante": str(real_visitante).strip(),
                    "goles_local_actual": ws.cell(r, 29).value,  # AC
                    "goles_visitante_actual": ws.cell(r, 30).value,  # AD
                }
            )

    wb.close()
    if not candidatos:
        raise SystemExit(f"No encuentro el partido: {local}-{visitante}")
    if len(candidatos) > 1:
        filas = ", ".join(str(c["fila"]) for c in candidatos)
        raise SystemExit(f"Partido ambiguo en filas: {filas}")
    return candidatos[0]


def applescript_text(s):
    return str(s).replace("\\", "\\\\").replace('"', '\\"')


def guardar_con_excel(ruta, fila, goles_local, goles_visitante):
    script = f'''
set workbookPath to POSIX file "{applescript_text(ruta)}"
tell application "Microsoft Excel"
    open workbook workbook file name workbookPath
    set wb to active workbook
    set value of range "AC{fila}" of worksheet "WORLDCUP" of wb to {goles_local}
    set value of range "AD{fila}" of worksheet "WORLDCUP" of wb to {goles_visitante}
    calculate full
    save wb
    close wb saving yes
end tell
'''
    subprocess.run(["osascript", "-e", script], check=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("args", nargs="+")
    parser.add_argument("--excel", default=DEFECTO_XLSX)
    parser.add_argument(
        "--force",
        action="store_true",
        help="Sobrescribe un resultado distinto ya escrito.",
    )
    ns = parser.parse_args()

    if len(ns.args) == 3:
        local, visitante = partir_codigo(ns.args[0])
        goles_local, goles_visitante = ns.args[1:]
    elif len(ns.args) == 4:
        local, visitante, goles_local, goles_visitante = ns.args
    else:
        raise SystemExit(
            "Uso: registrar_resultado.py 'Local-Visitante' GL GV "
            "o registrar_resultado.py Local Visitante GL GV"
        )

    try:
        goles_local = int(goles_local)
        goles_visitante = int(goles_visitante)
    except ValueError:
        raise SystemExit("Los goles deben ser numeros enteros.")
    if goles_local < 0 or goles_visitante < 0:
        raise SystemExit("Los goles no pueden ser negativos.")

    ruta = os.path.abspath(ns.excel)
    if not os.path.exists(ruta):
        raise SystemExit(f"No encuentro el Excel: {ruta}")

    partido = localizar_partido(ruta, local, visitante)
    actual = (partido["goles_local_actual"], partido["goles_visitante_actual"])
    nuevo = (goles_local, goles_visitante)
    if all(v not in (None, "") for v in actual) and tuple(map(int, actual)) != nuevo:
        if not ns.force:
            raise SystemExit(
                "Ese partido ya tiene otro resultado "
                f"({actual[0]}-{actual[1]}). Usa --force para sobrescribir."
            )

    guardar_con_excel(ruta, partido["fila"], goles_local, goles_visitante)
    print(
        "OK: "
        f"{partido['local']} {goles_local}-{goles_visitante} "
        f"{partido['visitante']} escrito en WORLDCUP fila {partido['fila']}."
    )


if __name__ == "__main__":
    main()
